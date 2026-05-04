"""
GPRD-Gold Trading Strategy — Full Python Replication + Rolling Sharpe Excel Sheet
"""
import math, warnings, shutil
from datetime import datetime, timedelta
from pathlib import Path
import numpy as np, pandas as pd
from scipy import stats
from scipy.stats import norm
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
warnings.filterwarnings("ignore")

def load_data(fp):
    g = pd.read_excel(fp, sheet_name="GPRD2008-Present")
    g.columns = ["Date_raw","DOW","GPRD","GPRD_Week_Avg","Date2"]
    g = g.dropna(subset=["GPRD"]); g["Date"] = pd.to_datetime(g["Date2"])
    g = g[["Date","DOW","GPRD"]].sort_values("Date").reset_index(drop=True)
    gold = pd.read_excel(fp, sheet_name="Gold_data")
    gold.columns = ["Date","Gold_Close"]; gold = gold.dropna(subset=["Gold_Close"])
    gold["Date"] = pd.to_datetime(gold["Date"]); gold = gold.sort_values("Date").reset_index(drop=True)
    sp = pd.read_excel(fp, sheet_name="S&P500").iloc[:,:3]
    sp.columns = ["Date","SP500_Level","SP500_1d_return"]
    sp = sp.dropna(subset=["SP500_Level"]); sp["Date"] = pd.to_datetime(sp["Date"])
    sp = sp.sort_values("Date").reset_index(drop=True)
    print(f"  GPRD: {len(g)} | Gold: {len(gold)} | S&P: {len(sp)}")
    return g, gold, sp

def weekly_gprd(gprd):
    d = gprd.copy(); d["WM"] = d["Date"] - pd.to_timedelta(d["Date"].dt.weekday, unit="D")
    return d.groupby("WM")["GPRD"].mean().to_dict()

def build_signals(gprd, dates, rw=504, buy_t=-0.2, sell_t=1.0):
    wk = weekly_gprd(gprd)
    gd = gprd["Date"].values; gv = gprd["GPRD"].values
    first_mon = dates.iloc[0] - timedelta(days=dates.iloc[0].weekday())
    recs = []
    for dt in dates:
        wm = dt - timedelta(days=dt.weekday())
        if wm == first_mon:
            recs.append({"Date":dt,"GPRD_Weekly":None,"Rolling_Mean":None,"Rolling_Std":None,"Z_score":None,"Signal":None}); continue
        pm = wm - timedelta(days=7); pwa = wk.get(pm)
        if pwa is None:
            recs.append({"Date":dt,"GPRD_Weekly":None,"Rolling_Mean":None,"Rolling_Std":None,"Z_score":None,"Signal":None}); continue
        pf = wm - timedelta(days=3); mask = gd <= np.datetime64(pf); na = mask.sum()
        if na < rw:
            recs.append({"Date":dt,"GPRD_Weekly":pwa,"Rolling_Mean":None,"Rolling_Std":None,"Z_score":None,"Signal":None}); continue
        w = gv[mask][-rw:]; rm = np.mean(w); rs = np.std(w, ddof=1)
        z = (pwa - rm)/rs if rs > 0 else 0.0
        sig = -1 if z > sell_t else (1 if z < buy_t else 0)
        recs.append({"Date":dt,"GPRD_Weekly":pwa,"Rolling_Mean":rm,"Rolling_Std":rs,"Z_score":z,"Signal":sig})
    return pd.DataFrame(recs)

def state_machine(signals):
    df = signals.copy()
    df["WM"] = df["Date"] - pd.to_timedelta(df["Date"].dt.weekday, unit="D")
    ws = df.dropna(subset=["Signal"]).drop_duplicates(subset=["WM"], keep="first").set_index("WM")["Signal"].to_dict()
    pos, positions = 0, []
    for _, r in df.iterrows():
        if pd.isna(r["Signal"]): positions.append(0); continue
        s = ws.get(r["WM"], 0)
        if s == 1: pos = 1
        elif s == -1: pos = 0
        positions.append(pos)
    df["Position"] = positions; return df

def backtest(gold, sp500, signals):
    df = gold.copy()
    df["Gold_1d_return"] = df["Gold_Close"].pct_change().shift(-1)
    df["Gold_1d_log_return"] = np.log(df["Gold_Close"]).diff().shift(-1)
    df["Gold_5d_log_return"] = np.log(df["Gold_Close"]).diff(5).shift(-5)
    df["Gold_10d_log_return"] = np.log(df["Gold_Close"]).diff(10).shift(-10)
    df = df.merge(sp500[["Date","SP500_Level"]], on="Date", how="left")
    df["SP500_1d_return"] = df["SP500_Level"].pct_change().shift(-1)
    df.loc[df["Gold_1d_return"].abs() > 0.5, "Gold_1d_return"] = np.nan
    df.loc[df["SP500_1d_return"].abs() > 0.5, "SP500_1d_return"] = np.nan
    sc = [c for c in ["Date","GPRD_Weekly","Rolling_Mean","Rolling_Std","Z_score","Signal","Position"] if c in signals.columns]
    df = df.merge(signals[sc], on="Date", how="left")
    df["Strategy_Return"] = df["Position"] * df["Gold_1d_return"]
    df["Strategy_Log_Return"] = df["Position"] * df["Gold_1d_log_return"]
    fi = df[df["Signal"].notna()].index[0]; df = df.loc[fi:].copy()
    df["Strategy_Value"] = 100*(1+df["Strategy_Return"].fillna(0)).cumprod()
    df["Gold_BH_Value"] = 100*(1+df["Gold_1d_return"].fillna(0)).cumprod()
    df["SP500_BH_Value"] = 100*(1+df["SP500_1d_return"].fillna(0)).cumprod()
    return df.reset_index(drop=True)

def metrics(res, rf=0.0274, label="Full"):
    rd = (1+rf)**(1/252)-1
    sr = res["Strategy_Return"].dropna(); gr = res["Gold_1d_return"].dropna()
    ny = len(res)/252
    st = res["Strategy_Value"].iloc[-1]/res["Strategy_Value"].iloc[0]
    gt = res["Gold_BH_Value"].iloc[-1]/res["Gold_BH_Value"].iloc[0]
    def mdd(v): return ((v - v.cummax())/v.cummax()).min()
    return {"Period":label,"Days":len(sr),"Years":round(ny,1),
        "Strat_Return":f"{(st-1)*100:.1f}%","Gold_Return":f"{(gt-1)*100:.1f}%",
        "Strat_CAGR":f"{(st**(1/ny)-1)*100:.2f}%","Gold_CAGR":f"{(gt**(1/ny)-1)*100:.2f}%",
        "Strat_Sharpe":round(((sr.mean()-rd)/sr.std())*np.sqrt(252),4),
        "Gold_Sharpe":round(((gr.mean()-rd)/gr.std())*np.sqrt(252),4),
        "Strat_MaxDD":f"{mdd(res['Strategy_Value'])*100:.2f}%",
        "Gold_MaxDD":f"{mdd(res['Gold_BH_Value'])*100:.2f}%",
        "In_Market":f"{(res['Position']==1).mean()*100:.1f}%",
        "Buy":(res["Signal"]==1).sum(),"Sell":(res["Signal"]==-1).sum(),"Hold":(res["Signal"]==0).sum(),
        "Strat_Final":round(res["Strategy_Value"].iloc[-1],2),
        "Gold_Final":round(res["Gold_BH_Value"].iloc[-1],2)}

def subperiod(res, s, e, l):
    sub = res[(res["Date"]>=s)&(res["Date"]<=e)].copy()
    if len(sub)==0: return None
    sub["Strategy_Value"]=100*(1+sub["Strategy_Return"].fillna(0)).cumprod()
    sub["Gold_BH_Value"]=100*(1+sub["Gold_1d_return"].fillna(0)).cumprod()
    sub["SP500_BH_Value"]=100*(1+sub["SP500_1d_return"].fillna(0)).cumprod()
    m = metrics(sub, label=l); m["SP500_Final"]=round(sub["SP500_BH_Value"].iloc[-1],2); return m

def stat_tests(res):
    df = res.dropna(subset=["Signal","Gold_5d_log_return"])
    sell = df.loc[df["Signal"]==-1,"Gold_5d_log_return"].values
    hold = df.loc[df["Signal"]==0,"Gold_5d_log_return"].values
    buy  = df.loc[df["Signal"]==1,"Gold_5d_log_return"].values
    def cd(a,b):
        sp=np.sqrt(((len(a)-1)*np.std(a,ddof=1)**2+(len(b)-1)*np.std(b,ddof=1)**2)/(len(a)+len(b)-2))
        return (np.mean(a)-np.mean(b))/sp if sp>0 else 0
    pairs=[("Buy vs Sell",buy,sell),("Buy vs Hold",buy,hold),("Sell vs Hold",sell,hold)]
    za=norm.ppf(0.975); zb=norm.ppf(0.80)
    print("\n"+"="*65+"\nSTATISTICAL VALIDATION (5-Day Forward Returns)\n"+"="*65)
    for l,d in [("Sell",sell),("Hold",hold),("Buy",buy)]:
        print(f"\n  {l}: n={len(d)}, mean={np.mean(d):.6f}, std={np.std(d,ddof=1):.6f}")
    print("\n"+"-"*65+"\nWELCH T-TESTS\n"+"-"*65)
    for l,a,b in pairs:
        t,p=stats.ttest_ind(a,b,equal_var=False)
        print(f"  {l}: t={t:.4f}, p={p:.6f}, sig@0.05={'Yes' if p<0.05 else 'No'}")
    print("\n"+"-"*65+"\nDIRECTIONAL ACCURACY & POWER (5-Day)\n"+"-"*65)
    print("  Tests whether each signal correctly predicts the direction of the 5-day forward return.")
    print("  Null: hit rate = 50% (no predictive power). One-tailed: Buy expects 5d>0, Sell expects 5d<0.")
    za1=norm.ppf(0.95)  # one-tailed alpha=0.05
    for l,d,sign,expect in [("Buy  (5d > 0)",buy,1,"positive"),("Sell (5d < 0)",sell,-1,"negative")]:
        n=len(d)
        if n==0: print(f"\n  {l}: no observations"); continue
        hits=int((d*sign>0).sum()); hr=hits/n
        se_hr=math.sqrt(hr*(1-hr)/n)
        z=(hr-0.5)/math.sqrt(0.25/n)
        p_val=1-norm.cdf(z)  # one-tailed: H1 is hr > 0.5
        ncp=(hr-0.5)/math.sqrt(0.25/n)
        power=1-norm.cdf(za1-ncp)
        n_req=math.ceil((za1+zb)**2*0.25/(hr-0.5)**2) if hr>0.5 else float('inf')
        print(f"\n  {l}:")
        print(f"    {expect} hits={hits}/{n}, rate={hr:.1%}, 95%CI=[{hr-1.96*se_hr:.1%}, {hr+1.96*se_hr:.1%}]")
        print(f"    z={z:.3f}, p={p_val:.4f} (one-tailed), sig@0.05={'Yes' if p_val<0.05 else 'No'}")
        print(f"    power={power:.1%}, n_needed={n_req}, have={n}")
    if len(hold)>0:
        t_h,p_h=stats.ttest_1samp(hold,0)
        pos_pct=(hold>0).mean(); neg_pct=(hold<0).mean()
        print(f"\n  Hold (expects ~0 mean 5d return):")
        print(f"    n={len(hold)}, mean={np.mean(hold):.6f}, std={np.std(hold,ddof=1):.6f}")
        print(f"    positive={pos_pct:.1%}, negative={neg_pct:.1%}")
        print(f"    t={t_h:.4f}, p={p_h:.4f}, sig@0.05={'Yes' if p_h<0.05 else 'No'}")
    print("\n"+"-"*65+"\nCOHEN'S D (Between Groups)\n"+"-"*65)
    for l,a,b in pairs:
        d=cd(a,b); mag="Neg" if abs(d)<0.2 else "Small" if abs(d)<0.5 else "Med" if abs(d)<0.8 else "Large"
        nn=math.ceil(((za+zb)/abs(d))**2) if abs(d)>0 else float('inf')
        print(f"  {l}: d={d:.4f} ({mag}), n_needed={nn}, have=({len(a)},{len(b)})")

def rolling_sharpe_sheet(fp, res, window=252, rf=0.0274):
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    rd = (1+rf)**(1/252)-1
    df = res[["Date","Strategy_Return","Gold_1d_return","SP500_1d_return"]].copy()
    for col,nm in [("Strategy_Return","Strat"),("Gold_1d_return","Gold"),("SP500_1d_return","SP500")]:
        ex = df[col]-rd; df[nm+"_Sharpe"] = (ex.rolling(window,min_periods=window).mean()/df[col].rolling(window,min_periods=window).std())*np.sqrt(252)
    sdf = df[["Date","Strat_Sharpe","Gold_Sharpe","SP500_Sharpe"]].dropna().reset_index(drop=True)
    print(f"  Rolling Sharpe: {len(sdf)} pts ({sdf['Date'].iloc[0].date()} to {sdf['Date'].iloc[-1].date()})")
    from openpyxl import Workbook
    dst=str(Path(fp).parent/"The_Finale_v2_output.xlsx")
    wb=Workbook(); sn="Rolling_Sharpe"
    ws=wb.active; ws.title=sn
    hf=Font(bold=True,size=11,name="Arial"); hfill=PatternFill("solid",fgColor="D9E1F2")
    for i,h in enumerate(["Date","Strategy (1Y Rolling)","Gold B&H (1Y Rolling)","S&P 500 B&H (1Y Rolling)"],1):
        c=ws.cell(row=1,column=i,value=h); c.font=hf; c.fill=hfill; c.alignment=Alignment(horizontal="center")
    for i,(_,r) in enumerate(sdf.iterrows(),2):
        ws.cell(row=i,column=1,value=r["Date"]).number_format="YYYY-MM-DD"
        ws.cell(row=i,column=2,value=r["Strat_Sharpe"]).number_format="0.0000"
        ws.cell(row=i,column=3,value=r["Gold_Sharpe"]).number_format="0.0000"
        ws.cell(row=i,column=4,value=r["SP500_Sharpe"]).number_format="0.0000"
    ws.column_dimensions["A"].width=14; ws.column_dimensions["B"].width=24
    ws.column_dimensions["C"].width=24; ws.column_dimensions["D"].width=28
    lr=len(sdf)+1
    ch=LineChart(); ch.title="Rolling 1-Year Sharpe Ratio"; ch.y_axis.title="Annualized Sharpe"
    ch.style=10; ch.width=30; ch.height=15
    dates=Reference(ws,min_col=1,min_row=2,max_row=lr)
    for ci,(nm,clr) in enumerate([("Strategy","2F5496"),("Gold B&H","C4953A"),("S&P 500 B&H","548235")],2):
        d=Reference(ws,min_col=ci,min_row=1,max_row=lr); ch.add_data(d,titles_from_data=True)
        ch.series[-1].graphicalProperties.line.width=18000; ch.series[-1].graphicalProperties.line.solidFill=clr
    ch.set_categories(dates); ch.x_axis.tickLblSkip=max(1,len(sdf)//10); ch.x_axis.numFmt="YYYY"
    ch.legend.position="b"; ws.add_chart(ch,"F2")
    wb.save(dst); return dst

def plot_charts(res, out_dir, rf=0.0274, window=252):
    rd = (1+rf)**(1/252)-1
    dates = pd.to_datetime(res["Date"])
    fig, axes = plt.subplots(3, 1, figsize=(14, 18))
    fig.suptitle("GPRD Gold Strategy Analysis", fontsize=15, fontweight="bold", y=0.99)

    # --- Chart 1: Rolling 1-Year Sharpe ---
    ax1 = axes[0]
    strat_ex = res["Strategy_Return"] - rd
    gold_ex  = res["Gold_1d_return"]  - rd
    strat_sharpe = (strat_ex.rolling(window, min_periods=window).mean() /
                    strat_ex.rolling(window, min_periods=window).std()) * np.sqrt(252)
    gold_sharpe  = (gold_ex.rolling(window, min_periods=window).mean() /
                    gold_ex.rolling(window, min_periods=window).std()) * np.sqrt(252)
    ax1.plot(dates, strat_sharpe, label="Strategy", color="#2F5496", linewidth=1.5)
    ax1.plot(dates, gold_sharpe,  label="Gold B&H",  color="#C4953A", linewidth=1.5)
    ax1.axhline(0, color="black", linewidth=0.8, linestyle="--", alpha=0.4)
    ax1.axhline(1, color="green", linewidth=0.8, linestyle=":",  alpha=0.4)
    ax1.set_title("Rolling 1-Year Sharpe Ratio", fontsize=12)
    ax1.set_ylabel("Annualized Sharpe")
    ax1.legend(); ax1.grid(True, alpha=0.3)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))

    # --- Chart 2: Cumulative Log Returns ---
    ax2 = axes[1]
    strat_cum = res["Strategy_Log_Return"].fillna(0).cumsum()
    gold_cum  = res["Gold_1d_log_return"].fillna(0).cumsum()
    ax2.plot(dates, strat_cum, label="Strategy", color="#2F5496", linewidth=1.5)
    ax2.plot(dates, gold_cum,  label="Gold B&H",  color="#C4953A", linewidth=1.5)
    ax2.axhline(0, color="black", linewidth=0.8, linestyle="--", alpha=0.4)
    ax2.set_title("Cumulative Log Returns: Strategy vs Gold", fontsize=12)
    ax2.set_ylabel("Cumulative Log Return")
    ax2.legend(); ax2.grid(True, alpha=0.3)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))

    # --- Chart 3: Strategy Value vs GPRD 504-Day MA ---
    ax3a = axes[2]
    ax3b = ax3a.twinx()
    gprd_ma = res["GPRD_Weekly"].rolling(504, min_periods=1).mean()
    ax3a.plot(dates, res["Strategy_Value"], label="Strategy Value", color="#2F5496", linewidth=1.5)
    ax3b.plot(dates, gprd_ma, label="GPRD 504-Day MA", color="#C4953A", linewidth=1.5, alpha=0.85)
    ax3a.set_title("Strategy Value vs GPRD 504-Day Moving Average", fontsize=12)
    ax3a.set_ylabel("Strategy Value (Base=100)", color="#2F5496")
    ax3b.set_ylabel("GPRD 504-Day MA",           color="#C4953A")
    ax3a.tick_params(axis="y", labelcolor="#2F5496")
    ax3b.tick_params(axis="y", labelcolor="#C4953A")
    ax3a.grid(True, alpha=0.3)
    ax3a.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    lines = ax3a.get_legend_handles_labels()[0] + ax3b.get_legend_handles_labels()[0]
    labels = ax3a.get_legend_handles_labels()[1] + ax3b.get_legend_handles_labels()[1]
    ax3a.legend(lines, labels, loc="upper left")

    plt.tight_layout()
    out_path = str(Path(out_dir)/"strategy_charts.png")
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {out_path}")
    return out_path

def main():
    fp=Path("/mnt/user-data/uploads/The_Finale_v2.xlsx")
    if not fp.exists(): fp=Path.home()/"Downloads"/"The_Finale_v2.xlsx"
    print("Loading data...")
    gprd,gold,sp500=load_data(str(fp))
    print("\nBuilding signals...")
    sigs=build_signals(gprd,gold["Date"])
    v=sigs.dropna(subset=["Signal"])
    print(f"  {len(v)} signal days: {v['Date'].iloc[0].date()} to {v['Date'].iloc[-1].date()}")
    print("\nState machine...")
    sigs=state_machine(sigs)
    ld=(sigs["Position"]==1).sum(); fd=(sigs["Position"]==0).sum()
    print(f"  Long:{ld} Flat:{fd} InMkt:{ld/(ld+fd)*100:.1f}%")
    print("\nBacktest...")
    res=backtest(gold,sp500,sigs)
    print("\n"+"="*65+"\nPERFORMANCE\n"+"="*65)
    m=metrics(res,label="Full Period (2010-2026)")
    print(f"\n  {m['Period']} ({m['Days']} days, {m['Years']} yrs)")
    for k in ["Strat_Return","Gold_Return","Strat_CAGR","Gold_CAGR","Strat_Sharpe","Gold_Sharpe","Strat_MaxDD","Gold_MaxDD","In_Market"]:
        print(f"    {k}: {m[k]}")
    print(f"    Signals: Buy={m['Buy']} Sell={m['Sell']} Hold={m['Hold']}")
    print(f"    Final: Strategy={m['Strat_Final']} Gold={m['Gold_Final']}")
    print("\n  IN-SAMPLE:")
    for s,e,l in [("2010-01-01","2019-12-31","2010-2019"),("2013-01-01","2016-12-31","2013-2016"),("2019-01-01","2022-12-31","2019-2022")]:
        sp=subperiod(res,s,e,l)
        if sp: print(f"    {l}: Strat={sp['Strat_Final']} Gold={sp['Gold_Final']} S&P={sp['SP500_Final']}")
    print("\n  OUT-OF-SAMPLE:")
    for s,e,l in [("2023-01-01","2026-04-30","2023-2026 (OOS)"),("2025-01-01","2026-04-30","2025-2026 (Fwd)")]:
        sp=subperiod(res,s,e,l)
        if sp: print(f"    {l}: Strat={sp['Strat_Final']} Gold={sp['Gold_Final']} S&P={sp['SP500_Final']}")
    stat_tests(res)
    print("\n"+"="*65+"\nCHARTS\n"+"="*65)
    plot_charts(res, fp.parent)
    print("\n"+"="*65+"\nROLLING SHARPE EXCEL SHEET\n"+"="*65)
    ox=rolling_sharpe_sheet(str(fp),res)
    csv_out=str(fp.parent/"strategy_results.csv")
    ec=[c for c in ["Date","Gold_Close","Gold_1d_return","Gold_1d_log_return","Gold_5d_log_return","Gold_10d_log_return","SP500_Level","GPRD_Weekly","Rolling_Mean","Rolling_Std","Z_score","Signal","Position","Strategy_Return","Strategy_Log_Return","Strategy_Value","Gold_BH_Value","SP500_BH_Value"] if c in res.columns]
    res[ec].to_csv(csv_out,index=False)
    xlsx_out=str(fp.parent/"The_Finale_v2_final.xlsx")
    shutil.copy2(ox,xlsx_out)
    print(f"\n  CSV: {csv_out}\n  XLSX: {xlsx_out}")

if __name__=="__main__": main()
