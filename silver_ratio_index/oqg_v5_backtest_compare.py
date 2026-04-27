"""
OQG v5 Backtest & Benchmark Comparison
=======================================
Place this file in the same directory as oqg_v5_improved.py.

Prompts for a start date, reruns the OQG model signal day-by-day from
that date to today, then compares performance against:
  - S&P 500 (SPY)
  - Gold ETF (GLD)
  - Silver ETF (SLV)
  - 60/40 Portfolio (60% SPY + 40% TLT)

Output: logs/portfolios/OQG_v5_Backtest_Report.xlsx
        4 sheets: Summary Statistics, Performance Comparison,
                  Daily Trade Log, Drawdown Analysis
"""

import os
import sys
import warnings
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

warnings.filterwarnings("ignore")

# ── Import from sibling module ────────────────────────────────────────────────
try:
    from oqg_v5_improved import (
        fetch_data, train_model_walkforward, get_signal,
        FEATURES, STARTING_CAPITAL, SLIPPAGE_FEE,
        REBALANCE_COOLDOWN, CONFIDENCE_MIN, MODEL_DIR,
    )
except ImportError as e:
    print(f"ERROR: Could not import from oqg_v5_improved.py\n  {e}")
    print("Make sure this file is in the same directory as oqg_v5_improved.py")
    sys.exit(1)

OUTPUT_FILE = "logs/portfolios/OQG_v5_Backtest_Report.xlsx"
os.makedirs("logs/portfolios", exist_ok=True)


# ── Style constants ───────────────────────────────────────────────────────────
NAVY        = "1F3864"
GOLD_C      = "B8860B"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
DARK_GRAY   = "595959"
GREEN_BEST  = "C6EFCE"
GREEN_TXT   = "375623"
RED_LIGHT   = "FFCCCC"
ORANGE_LIGHT= "FFE6CC"


# ── Style helpers ─────────────────────────────────────────────────────────────
def _thin():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, col, val, bg=NAVY, fg=WHITE, wrap=False, size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color=fg, size=size, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def _cell(ws, row, col, v, fmt=None, bold=False, bg=None, center=True):
    c = ws.cell(row=row, column=col, value=v)
    c.font      = Font(bold=bold, size=10, name="Arial")
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    c.border    = _thin()
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _border_range(ws, r1, r2, c1, c2):
    b = _thin()
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = b

def _title_row(ws, text, merge_to, row=1, bg=NAVY, size=13, height=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, size=size, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height
    return c


# ── Prompt ────────────────────────────────────────────────────────────────────
def prompt_start_date():
    while True:
        raw = input("\nEnter backtest start date (YYYY-MM-DD): ").strip()
        try:
            d = datetime.strptime(raw, "%Y-%m-%d")
            if d >= datetime.now():
                print("  Date must be in the past. Try again.")
                continue
            if (datetime.now() - d).days < 60:
                print("  Warning: very short window — fewer than 60 days back.")
            return d
        except ValueError:
            print("  Invalid format. Use YYYY-MM-DD  e.g. 2020-01-01")


# ── OQG backtest ──────────────────────────────────────────────────────────────
def run_oqg_backtest(df, start_date, model, scaler):
    subset = df[df.index >= pd.to_datetime(start_date)].copy()
    if subset.empty:
        raise ValueError("No data from the given start date.")

    cash             = STARTING_CAPITAL
    units_held       = 0.0
    current_position = "CASH"
    days_in_pos      = 0
    prev_equity      = STARTING_CAPITAL
    rows             = []

    for date_idx, row in subset.iterrows():
        target, ml_pred, ml_prob = get_signal(row, model, scaler)
        fee_paid    = 0.0
        rotated     = False
        days_in_pos += 1

        can_rotate = (days_in_pos >= REBALANCE_COOLDOWN) or (current_position == "CASH")
        if current_position != target and can_rotate:
            rotated = True
            if current_position == "SILVER":
                cash += units_held * row["Silver"]
            elif current_position == "GOLD":
                cash += units_held * row["Gold"]
            units_held = 0.0

            if target != "CASH":
                fee_paid   = cash * SLIPPAGE_FEE
                cash      -= fee_paid
                buy_price  = row["Silver"] if target == "SILVER" else row["Gold"]
                units_held = cash / buy_price
                cash       = 0.0

            current_position = target
            days_in_pos      = 0

        if current_position == "CASH":
            total_equity = cash
        elif current_position == "SILVER":
            total_equity = units_held * row["Silver"]
        else:
            total_equity = units_held * row["Gold"]

        day_ret = (total_equity - prev_equity) / prev_equity if prev_equity > 0 else 0.0
        cum_ret = (total_equity - STARTING_CAPITAL) / STARTING_CAPITAL
        prev_equity = total_equity

        rows.append({
            "Date":           date_idx,
            "Position":       current_position,
            "Signal":         target,
            "Rotated":        "YES" if rotated else "",
            "ML_Pred":        ml_pred,
            "Silver_Prob":    round(ml_prob, 4),
            "Confidence_Met": "YES" if ml_prob >= CONFIDENCE_MIN else "NO",
            "Gold_Price":     round(row["Gold"], 2),
            "Silver_Price":   round(row["Silver"], 4),
            "SPY_Ret_20d":    round(float(row.get("SPY_Ret_20d", 0)), 4),
            "VIX_Level":      round(float(row.get("VIX_Level", 1.0)) * 20, 2),
            "Gold_MA200":     round(row["Gold_MA200"], 2),
            "Silver_HighVol": bool(row["Silver_HighVol"]),
            "Gold_AboveMA":   bool(row["Gold_AboveMA"]),
            "Units_Held":     round(units_held, 6),
            "Cash":           round(cash, 2),
            "Total_Equity":   round(total_equity, 2),
            "Day_Return_Pct": round(day_ret * 100, 4),
            "Cum_Return_Pct": round(cum_ret * 100, 4),
            "Fee_Paid":       round(fee_paid, 2),
        })

    return pd.DataFrame(rows)


# ── Benchmark equity curves ───────────────────────────────────────────────────
def build_benchmarks(start_date):
    s = pd.to_datetime(start_date).strftime("%Y-%m-%d")
    e = datetime.now().strftime("%Y-%m-%d")
    raw    = yf.download(["SPY", "GLD", "SLV", "TLT"], start=s, end=e, progress=False)
    prices = (raw["Close"].copy()
              if isinstance(raw.columns, pd.MultiIndex)
              else raw.copy())
    prices = prices.ffill().bfill()

    bm = pd.DataFrame(index=prices.index)
    for ticker in ["SPY", "GLD", "SLV", "TLT"]:
        if ticker in prices.columns:
            bm[ticker] = (prices[ticker] / prices[ticker].iloc[0] * STARTING_CAPITAL).round(2)

    if "SPY" in prices.columns and "TLT" in prices.columns:
        bm["60_40"] = (
            (0.60 * prices["SPY"] / prices["SPY"].iloc[0] +
             0.40 * prices["TLT"] / prices["TLT"].iloc[0])
            * STARTING_CAPITAL
        ).round(2)

    return bm, prices


# ── Statistics ────────────────────────────────────────────────────────────────
def compute_stats(equity, label="Strategy", spy_returns=None):
    equity = equity.dropna()
    if len(equity) < 5:
        return {}
    daily  = equity.pct_change().dropna()
    years  = len(equity) / 252

    total_ret  = (equity.iloc[-1] - equity.iloc[0]) / equity.iloc[0]
    cagr       = (1 + total_ret) ** (1 / years) - 1 if years > 0 else 0.0
    ann_vol    = daily.std() * np.sqrt(252)
    sharpe     = daily.mean() / daily.std() * np.sqrt(252) if daily.std() > 0 else 0.0
    downside   = daily[daily < 0].std() * np.sqrt(252)
    sortino    = daily.mean() * 252 / downside if downside > 0 else 0.0
    cum_max    = equity.cummax()
    dd         = (equity - cum_max) / cum_max
    max_dd     = dd.min()
    calmar     = cagr / abs(max_dd) if max_dd != 0 else 0.0
    win_rate   = (daily > 0).sum() / len(daily)
    best_day   = daily.max()
    worst_day  = daily.min()

    # Max drawdown duration (days)
    dd_dur, cur = 0, 0
    for v in dd < 0:
        cur = cur + 1 if v else 0
        dd_dur = max(dd_dur, cur)

    # Beta / Alpha
    beta, alpha = np.nan, np.nan
    if spy_returns is not None:
        aligned = spy_returns.reindex(daily.index).dropna()
        strat   = daily.reindex(aligned.index).dropna()
        if len(aligned) > 10 and aligned.std() > 0:
            cov   = np.cov(strat, aligned)
            beta  = cov[0, 1] / cov[1, 1]
            alpha = (strat.mean() - beta * aligned.mean()) * 252

    return {
        "Strategy":               label,
        "Total Return":           total_ret,
        "CAGR":                   cagr,
        "Ann. Volatility":        ann_vol,
        "Sharpe Ratio":           sharpe,
        "Sortino Ratio":          sortino,
        "Max Drawdown":           max_dd,
        "Calmar Ratio":           calmar,
        "Max DD Duration (days)": dd_dur,
        "Win Rate":               win_rate,
        "Best Day":               best_day,
        "Worst Day":              worst_day,
        "Beta vs SPY":            beta,
        "Alpha vs SPY (ann)":     alpha,
        "Trading Days":           len(equity),
    }


# ── Sheet 1: Summary Statistics ───────────────────────────────────────────────
def write_summary_stats(wb, oqg_df, benchmarks, start_date):
    ws = wb.create_sheet("Summary Statistics")
    ws.sheet_view.showGridLines = False

    # Title
    date_range = (f"{pd.to_datetime(start_date).strftime('%b %d, %Y')} "
                  f"→ {datetime.now().strftime('%b %d, %Y')}")
    _title_row(ws, f"OQG v5 Improved — Strategy Performance Report  |  {date_range}",
               merge_to=7, row=1, size=12)

    # Build equity series for each strategy
    oqg_eq = oqg_df.set_index("Date")["Total_Equity"]
    oqg_eq.index = pd.to_datetime(oqg_eq.index)

    spy_daily = None
    if "SPY" in benchmarks.columns:
        spy_daily = benchmarks["SPY"].pct_change().dropna()

    strategies = {
        "OQG v5 Model":     oqg_eq,
        "S&P 500 (SPY)":    benchmarks.get("SPY"),
        "Gold ETF (GLD)":   benchmarks.get("GLD"),
        "Silver ETF (SLV)": benchmarks.get("SLV"),
        "60/40 Portfolio":  benchmarks.get("60_40"),
    }

    all_stats = []
    for label, eq in strategies.items():
        if eq is None or (hasattr(eq, "empty") and eq.empty):
            continue
        eq = eq.dropna()
        if len(eq) >= 5:
            all_stats.append(compute_stats(eq, label=label, spy_returns=spy_daily))

    if not all_stats:
        ws.cell(row=3, column=1, value="Insufficient data for statistics.")
        return ws

    n_strats = len(all_stats)

    # Column headers
    strat_colors = ["1F3864", "C55A11", "7F6000", "595959", "375623"]
    _hdr(ws, 3, 1, "Metric", bg="2F5496", size=10)
    for ci, (s, color) in enumerate(zip(all_stats, strat_colors), 2):
        c = _hdr(ws, 3, ci, s["Strategy"], bg=color, wrap=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    # Metric rows: (display label, dict key, number format, higher_is_better)
    metrics = [
        ("Total Return",           "Total Return",           "+0.00%;-0.00%;-",  True),
        ("CAGR (Annualized)",       "CAGR",                   "+0.00%;-0.00%;-",  True),
        ("Annualized Volatility",   "Ann. Volatility",        "0.00%",            False),
        ("Sharpe Ratio",            "Sharpe Ratio",           "0.00",             True),
        ("Sortino Ratio",           "Sortino Ratio",          "0.00",             True),
        ("Max Drawdown",            "Max Drawdown",           "0.00%;-0.00%;-",   True),
        ("Calmar Ratio",            "Calmar Ratio",           "0.00",             True),
        ("Max DD Duration (days)",  "Max DD Duration (days)", "#,##0",            False),
        ("Win Rate (% days +)",     "Win Rate",               "0.00%",            True),
        ("Best Single Day",         "Best Day",               "+0.00%;-0.00%;-",  True),
        ("Worst Single Day",        "Worst Day",              "+0.00%;-0.00%;-",  True),
        ("Beta vs S&P 500",         "Beta vs SPY",            "0.00",             False),
        ("Alpha vs S&P 500 (ann)",  "Alpha vs SPY (ann)",     "+0.00%;-0.00%;-",  True),
        ("Trading Days",            "Trading Days",           "#,##0",            False),
    ]

    row = 4
    for mi, (lbl, key, fmt, higher_better) in enumerate(metrics):
        alt_bg = LIGHT_GRAY if mi % 2 == 0 else WHITE

        # Label cell
        c = ws.cell(row=row, column=1, value=lbl)
        c.font      = Font(bold=True, size=10, name="Arial")
        c.fill      = PatternFill("solid", fgColor="DCE6F1")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _thin()

        raw_vals = [s.get(key, np.nan) for s in all_stats]
        nums     = [v for v in raw_vals if pd.notna(v) and isinstance(v, (int, float))]
        best     = (max(nums) if higher_better else min(nums)) if nums else None

        for ci, v in enumerate(raw_vals, 2):
            if pd.notna(v) and isinstance(v, (int, float)):
                c = _cell(ws, row, ci, v, fmt=fmt,
                          bold=(v == best),
                          bg=GREEN_BEST if v == best else alt_bg)
                if v == best:
                    c.font = Font(bold=True, size=10, name="Arial", color=GREEN_TXT)
            else:
                _cell(ws, row, ci, "N/A", bg=alt_bg)
        row += 1

    # ── OQG-only detail block ─────────────────────────────────────────────────
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_strats + 1)
    c = ws.cell(row=row, column=1, value="OQG v5 Model — Additional Details")
    c.font      = Font(bold=True, color=WHITE, size=10, name="Arial")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center")
    row += 1

    oqg_extras = [
        ("Total Rotations",    (oqg_df["Rotated"] == "YES").sum(), "#,##0",       None),
        ("Total Fees Paid",    oqg_df["Fee_Paid"].sum(),           "$#,##0.00",   None),
        ("GOLD Days",          (oqg_df["Position"] == "GOLD").sum(),"#,##0",      "FFF3CD"),
        ("SILVER Days",        (oqg_df["Position"] == "SILVER").sum(),"#,##0",    "E8E8E8"),
        ("CASH Days",          (oqg_df["Position"] == "CASH").sum(), "#,##0",     "D4EDDA"),
        ("Starting Capital",   STARTING_CAPITAL,                   "$#,##0.00",   None),
        ("Final Equity",       oqg_df["Total_Equity"].iloc[-1],    "$#,##0.00",   None),
    ]
    for i, (lbl, v, fmt, bg) in enumerate(oqg_extras):
        alt = LIGHT_GRAY if i % 2 == 0 else WHITE
        c = ws.cell(row=row, column=1, value=lbl)
        c.font   = Font(bold=True, size=10, name="Arial")
        c.fill   = PatternFill("solid", fgColor="DCE6F1")
        c.border = _thin()
        c.alignment = Alignment(horizontal="left", vertical="center")
        _cell(ws, row, 2, v, fmt=fmt, bold=True, bg=bg if bg else alt)
        row += 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_strats + 1)
    c = ws.cell(row=row, column=1, value="Notes")
    c.font = Font(bold=True, color=WHITE, size=10, name="Arial")
    c.fill = PatternFill("solid", fgColor=DARK_GRAY)
    c.alignment = Alignment(horizontal="center")
    row += 1
    notes = [
        f"Starting capital for all strategies: ${STARTING_CAPITAL:,.0f}",
        f"OQG slippage fee: {SLIPPAGE_FEE * 100:.1f}% applied per rotation  |  "
        f"Rebalance cooldown: {REBALANCE_COOLDOWN} days  |  "
        f"Confidence threshold: {CONFIDENCE_MIN * 100:.0f}%",
        "Benchmarks (SPY, GLD, SLV, 60/40) are buy-and-hold with NO transaction costs",
        "Beta and Alpha computed vs S&P 500 (SPY) daily returns",
        "Sharpe and Sortino assume risk-free rate of 0%",
        "Green cell = best value in that row across all strategies",
    ]
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_strats + 1)
        c = ws.cell(row=row, column=1, value=f"• {note}")
        c.font = Font(size=9, italic=True, name="Arial")
        c.fill = PatternFill("solid", fgColor="F8F8F8")
        row += 1

    _set_widths(ws, [28] + [18] * n_strats)
    ws.freeze_panes = "B4"
    return ws


# ── Sheet 2: Performance Comparison (equity curves + chart) ───────────────────
def write_performance_comparison(wb, oqg_df, benchmarks):
    ws = wb.create_sheet("Performance Comparison")
    ws.sheet_view.showGridLines = False

    _title_row(ws, "OQG v5 — Daily Equity Curves vs Benchmarks", merge_to=8)

    cols = ["Date", "OQG v5 ($)", "S&P 500 ($)", "Gold ETF ($)",
            "Silver ETF ($)", "60/40 ($)", "OQG Cum %", "SPY Cum %"]
    for ci, h in enumerate(cols, 1):
        _hdr(ws, 2, ci, h, wrap=True)
    ws.row_dimensions[2].height = 28

    # Align OQG onto benchmark index
    oqg_eq = oqg_df.set_index("Date")["Total_Equity"]
    oqg_eq.index = pd.to_datetime(oqg_eq.index)
    combined = benchmarks.copy()
    combined["OQG"] = oqg_eq.reindex(combined.index).ffill()
    combined = combined.dropna(subset=["OQG"])

    alt = PatternFill("solid", fgColor=LIGHT_GRAY)
    usd = "$#,##0.00"
    pct = "+0.00%;-0.00%;-"

    for ri, (dt, row) in enumerate(combined.iterrows(), 3):
        rf = alt if ri % 2 == 0 else None

        def _v(key):
            v = row.get(key)
            return None if (v is None or (isinstance(v, float) and np.isnan(v))) else v

        oqg_v = _v("OQG")
        spy_v = _v("SPY")
        oqg_cum = (oqg_v - STARTING_CAPITAL) / STARTING_CAPITAL if oqg_v else None
        spy_cum = (spy_v - STARTING_CAPITAL) / STARTING_CAPITAL if spy_v else None

        row_data = [
            (dt.strftime("%Y-%m-%d"), None),
            (oqg_v, usd),
            (spy_v, usd),
            (_v("GLD"), usd),
            (_v("SLV"), usd),
            (_v("60_40"), usd),
            (oqg_cum, pct),
            (spy_cum, pct),
        ]
        for ci, (v, fmt) in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(size=9, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt and v is not None:
                c.number_format = fmt
            if rf:
                c.fill = rf

    _set_widths(ws, [13, 14, 14, 14, 14, 14, 13, 13])
    ws.freeze_panes = "A3"
    _border_range(ws, 2, ws.max_row, 1, len(cols))

    # Chart
    max_row = ws.max_row
    if max_row > 4:
        chart = LineChart()
        chart.title  = "Portfolio Growth — OQG v5 vs Benchmarks"
        chart.style  = 10
        chart.height = 15
        chart.width  = 28
        chart.y_axis.title = "Portfolio Value ($)"
        chart.x_axis.title = "Date"

        colors  = ["1F3864", "FF6600", "B8860B", "808080", "2E8B57"]
        widths  = [25000,    15000,    15000,    15000,    15000]
        for idx, (col_i, color, lw) in enumerate(zip([2, 3, 4, 5, 6], colors, widths)):
            ref = Reference(ws, min_col=col_i, min_row=2, max_row=max_row)
            chart.add_data(ref, titles_from_data=True)
            chart.series[idx].graphicalProperties.line.solidFill = color
            chart.series[idx].graphicalProperties.line.width = lw

        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=max_row))
        ws.add_chart(chart, "J3")

    return ws


# ── Sheet 3: Daily Trade Log ──────────────────────────────────────────────────
def write_daily_log(wb, oqg_df):
    ws = wb.create_sheet("Daily Trade Log")

    _title_row(ws, "OQG v5 Improved — Daily Trade Log", merge_to=20)

    headers = [
        "Date", "Position", "Signal", "Rotated", "ML Pred",
        "Silver Prob", "Conf. Met", "Gold Price ($)", "Silver Price ($)",
        "SPY 20d Ret", "VIX Level", "Gold MA200 ($)",
        "Silver HighVol", "Gold Above MA", "Units Held",
        "Cash ($)", "Total Equity ($)", "Day Return %", "Cum Return %", "Fee Paid ($)",
    ]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 2, ci, h, wrap=True)
    ws.row_dimensions[2].height = 30

    gold_fill   = PatternFill("solid", fgColor="FFF3CD")
    silver_fill = PatternFill("solid", fgColor="E8E8E8")
    cash_fill   = PatternFill("solid", fgColor="D4EDDA")
    alt_fill    = PatternFill("solid", fgColor=LIGHT_GRAY)
    usd = "$#,##0.00"

    for ri, (_, row) in enumerate(oqg_df.iterrows(), 3):
        pos = row["Position"]
        rf  = (gold_fill if pos == "GOLD" else
               silver_fill if pos == "SILVER" else
               cash_fill if pos == "CASH" else
               (alt_fill if ri % 2 == 0 else None))

        dt_val = row["Date"]
        if hasattr(dt_val, "strftime"):
            dt_val = dt_val.strftime("%Y-%m-%d")

        cells = [
            (dt_val,                                     None),
            (row["Position"],                            None),
            (row["Signal"],                              None),
            (row["Rotated"],                             None),
            (int(row["ML_Pred"]),                        None),
            (row["Silver_Prob"],                         "0.0000"),
            (row["Confidence_Met"],                      None),
            (row["Gold_Price"],                          usd),
            (row["Silver_Price"],                        "$#,##0.0000"),
            (round(row["SPY_Ret_20d"], 4),               "0.00%"),
            (row["VIX_Level"],                           "0.00"),
            (row["Gold_MA200"],                          usd),
            ("Yes" if row["Silver_HighVol"] else "No",   None),
            ("Yes" if row["Gold_AboveMA"]   else "No",   None),
            (row["Units_Held"],                          "0.000000"),
            (row["Cash"],                                usd),
            (row["Total_Equity"],                        usd),
            (row["Day_Return_Pct"] / 100,                "+0.00%;-0.00%;-"),
            (row["Cum_Return_Pct"] / 100,                "+0.00%;-0.00%;-"),
            (row["Fee_Paid"],                            usd),
        ]
        for ci, (v, fmt) in enumerate(cells, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(size=9, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt:
                c.number_format = fmt
            if rf:
                c.fill = rf

        # Highlight rotations in the Rotated column
        if row["Rotated"] == "YES":
            ws.cell(row=ri, column=4).font = Font(bold=True, color="C00000", size=9, name="Arial")

    _set_widths(ws, [13,9,9,9,8,11,11,14,14,12,10,14,13,13,12,14,16,12,13,12])
    ws.freeze_panes = "A3"
    _border_range(ws, 2, ws.max_row, 1, len(headers))

    # Quick stats sidebar
    sc = len(headers) + 2
    _hdr(ws, 2, sc, "Quick Stats", bg=NAVY)
    ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=sc + 1)
    items = [
        ("Total Days",    len(oqg_df)),
        ("Rotations",     (oqg_df["Rotated"] == "YES").sum()),
        ("Total Fees",    f"${oqg_df['Fee_Paid'].sum():,.2f}"),
        ("Final Equity",  f"${oqg_df['Total_Equity'].iloc[-1]:,.2f}"),
        ("Cum Return",    f"{oqg_df['Cum_Return_Pct'].iloc[-1]:.2f}%"),
        ("GOLD Days",     (oqg_df["Position"] == "GOLD").sum()),
        ("SILVER Days",   (oqg_df["Position"] == "SILVER").sum()),
        ("CASH Days",     (oqg_df["Position"] == "CASH").sum()),
    ]
    for i, (k, v) in enumerate(items, 3):
        ws.cell(row=i, column=sc,     value=k).font = Font(bold=True, size=9, name="Arial")
        ws.cell(row=i, column=sc + 1, value=v).font = Font(size=9,          name="Arial")

    return ws


# ── Sheet 4: Drawdown Analysis ────────────────────────────────────────────────
def write_drawdown_sheet(wb, oqg_df, benchmarks):
    ws = wb.create_sheet("Drawdown Analysis")
    ws.sheet_view.showGridLines = False

    _title_row(ws, "Drawdown Analysis — OQG v5 vs Benchmarks", merge_to=7)

    headers = ["Date", "OQG DD %", "SPY DD %", "GLD DD %", "SLV DD %", "60/40 DD %", "OQG Position"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 2, ci, h, wrap=True)
    ws.row_dimensions[2].height = 28

    # Build drawdown series
    oqg_eq = oqg_df.set_index("Date")["Total_Equity"]
    oqg_eq.index = pd.to_datetime(oqg_eq.index)
    oqg_dd = (oqg_eq - oqg_eq.cummax()) / oqg_eq.cummax()

    dd = {"OQG": oqg_dd}
    for key in ["SPY", "GLD", "SLV", "60_40"]:
        if key in benchmarks.columns:
            s = benchmarks[key].dropna()
            dd[key] = (s - s.cummax()) / s.cummax()

    oqg_pos = oqg_df.set_index("Date")["Position"]
    oqg_pos.index = pd.to_datetime(oqg_pos.index)
    alt = PatternFill("solid", fgColor=LIGHT_GRAY)
    pct = "0.00%;-0.00%;-"

    for ri, dt in enumerate(oqg_dd.index, 3):
        rf = alt if ri % 2 == 0 else None

        def _get_dd(key):
            s = dd.get(key)
            if s is None:
                return None
            try:
                v = float(s.loc[dt])
                return None if np.isnan(v) else v
            except KeyError:
                return None

        row_data = [
            (dt.strftime("%Y-%m-%d"), None),
            (_get_dd("OQG"),   pct),
            (_get_dd("SPY"),   pct),
            (_get_dd("GLD"),   pct),
            (_get_dd("SLV"),   pct),
            (_get_dd("60_40"), pct),
            (oqg_pos.get(dt, ""), None),
        ]
        for ci, (v, fmt) in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(size=9, name="Arial")
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt and v is not None:
                c.number_format = fmt
            # Color-code severity
            if isinstance(v, float):
                if v < -0.20:
                    c.fill = PatternFill("solid", fgColor="FFCCCC")   # Deep red
                elif v < -0.10:
                    c.fill = PatternFill("solid", fgColor="FFE6CC")   # Orange
                elif v < -0.05:
                    c.fill = PatternFill("solid", fgColor="FFFACD")   # Yellow
                elif rf:
                    c.fill = rf
            elif rf:
                c.fill = rf

    _set_widths(ws, [13, 13, 13, 13, 13, 13, 13])
    ws.freeze_panes = "A3"
    _border_range(ws, 2, ws.max_row, 1, len(headers))

    # Drawdown chart
    max_row = ws.max_row
    if max_row > 4:
        chart = LineChart()
        chart.title  = "Drawdown Over Time"
        chart.style  = 10
        chart.height = 15
        chart.width  = 28
        chart.y_axis.title = "Drawdown (%)"
        chart.x_axis.title = "Date"

        colors = ["1F3864", "FF6600", "B8860B", "808080", "2E8B57"]
        for idx, (col_i, color) in enumerate(zip([2, 3, 4, 5, 6], colors)):
            ref = Reference(ws, min_col=col_i, min_row=2, max_row=max_row)
            chart.add_data(ref, titles_from_data=True)
            chart.series[idx].graphicalProperties.line.solidFill = color
            chart.series[idx].graphicalProperties.line.width = 20000 if idx == 0 else 15000

        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=max_row))
        ws.add_chart(chart, "I3")

        # Legend note
        ws.cell(row=3, column=9,
                value="🔴 > 20% drawdown  |  🟠 10–20%  |  🟡 5–10%").font = Font(
                size=9, italic=True, name="Arial")

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("  OQG v5 Improved — Backtest & Benchmark Comparison")
    print("=" * 70)

    start_date = prompt_start_date()
    lookback   = max(1500, (datetime.now() - start_date).days + 700)

    print(f"\n[1/5] Fetching data  (lookback: {lookback} days)...")
    df       = fetch_data(lookback_days=lookback)
    df_slice = df[df.index >= pd.to_datetime(start_date)]
    print(f"      {len(df_slice)} trading days in backtest window  "
          f"({df_slice.index.min().date()} → {df_slice.index.max().date()})")

    print("\n[2/5] Loading / training model...")
    model, scaler, metrics = train_model_walkforward(df)
    print(f"      Model AUC: {metrics['auc']:.3f}")

    print("\n[3/5] Running OQG backtest...")
    oqg_df = run_oqg_backtest(df, start_date, model, scaler)
    rotations = (oqg_df["Rotated"] == "YES").sum()
    final_eq  = oqg_df["Total_Equity"].iloc[-1]
    print(f"      {len(oqg_df)} days  |  {rotations} rotations  |  "
          f"Final equity: ${final_eq:,.2f}")

    print("\n[4/5] Fetching benchmarks (SPY, GLD, SLV, TLT)...")
    benchmarks, _ = build_benchmarks(start_date)
    print(f"      {len(benchmarks.columns)} series loaded")

    print("\n[5/5] Building Excel report...")
    wb = Workbook()
    wb.remove(wb.active)   # Remove default blank sheet

    write_summary_stats(wb, oqg_df, benchmarks, start_date)
    write_performance_comparison(wb, oqg_df, benchmarks)
    write_daily_log(wb, oqg_df)
    write_drawdown_sheet(wb, oqg_df, benchmarks)

    wb.save(OUTPUT_FILE)
    print(f"\n  ✓  Report saved → {OUTPUT_FILE}")

    # ── Console summary table ─────────────────────────────────────────────────
    oqg_eq = oqg_df.set_index("Date")["Total_Equity"]
    oqg_eq.index = pd.to_datetime(oqg_eq.index)
    spy_eq = benchmarks.get("SPY")
    spy_ret = spy_eq.pct_change().dropna() if spy_eq is not None else None

    oqg_s = compute_stats(oqg_eq, spy_returns=spy_ret)
    spy_s = compute_stats(spy_eq) if spy_eq is not None else {}

    def fp(v):
        return f"{v * 100:+.2f}%" if pd.notna(v) else "N/A"
    def f2(v):
        return f"{v:.3f}" if pd.notna(v) else "N/A"

    print("\n" + "=" * 60)
    print(f"  {'Metric':<26} {'OQG v5':>12}  {'S&P 500':>12}")
    print("  " + "-" * 54)
    rows_console = [
        ("Total Return",      fp(oqg_s.get("Total Return")),   fp(spy_s.get("Total Return"))),
        ("CAGR",              fp(oqg_s.get("CAGR")),           fp(spy_s.get("CAGR"))),
        ("Sharpe Ratio",      f2(oqg_s.get("Sharpe Ratio")),   f2(spy_s.get("Sharpe Ratio"))),
        ("Sortino Ratio",     f2(oqg_s.get("Sortino Ratio")),  f2(spy_s.get("Sortino Ratio"))),
        ("Max Drawdown",      fp(oqg_s.get("Max Drawdown")),   fp(spy_s.get("Max Drawdown"))),
        ("Ann. Volatility",   fp(oqg_s.get("Ann. Volatility")),fp(spy_s.get("Ann. Volatility"))),
        ("Win Rate",          fp(oqg_s.get("Win Rate")),        fp(spy_s.get("Win Rate"))),
        ("Beta vs SPY",       f2(oqg_s.get("Beta vs SPY")),    "1.000"),
    ]
    for lbl, oqg_v, spy_v in rows_console:
        print(f"  {lbl:<26} {oqg_v:>12}  {spy_v:>12}")
    print("=" * 60)
    print(f"  Full report: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    main()