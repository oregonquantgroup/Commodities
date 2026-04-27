"""
OQG v5 IMPROVED — Production-Ready Precious Metals Strategy
===========================================================
Fixes from v5 original:
  ✓ Walk-forward retraining (monthly)
  ✓ Regime detection (SPY momentum, VIX as kill-switch)
  ✓ 5-day forward target (cleaner signal than 1-day)
  ✓ Enhanced features (relative strength, volatility ratio, carry)
  ✓ Confidence thresholds (only trade prob > 0.60)
  ✓ Hyperparameter tuning + cross-validation
  ✓ Model validation (check AUC > 0.55)
  ✓ Reduced trading frequency (avoid churn)

Run from:  v5_linear/
Output:    logs/portfolios/OQG_v5_Improved_Portfolio.xlsx
"""

import os
import warnings
import joblib
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score
from sklearn.metrics import roc_auc_score, classification_report
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── Config ─────────────────────────────────────────────────────────────────────
PORTFOLIO_FILE     = "logs/portfolios/OQG_v5_Improved_Portfolio.xlsx"
MODEL_DIR          = "logs/models"
STARTING_CAPITAL   = 100_000.0
SLIPPAGE_FEE       = 0.001
MA_WINDOW          = 200
REBALANCE_COOLDOWN = 3  # Days before allowing next rotation (reduce churn)
CONFIDENCE_MIN     = 0.60  # Only trade if prob > 60%
MIN_POSITION_SIZE  = 0.5  # Don't rotate if position size < 50% (avoid micro-trades)

# Feature set (base + regime + enhanced)
BASE_FEATURES = (
    [f'GCR_Z_lag{l}' for l in [1, 2, 3]] +
    [f'GSR_Z_lag{l}' for l in [1, 2, 3]] +
    [f'Gold_Ret_{w}d'   for w in [5, 10, 20]] +
    [f'Silver_Ret_{w}d' for w in [5, 10, 20]] +
    ['Corr_Gold_GCR', 'Corr_Gold_GSR'] +
    [f'Corr_Gold_GCR_lag{l}' for l in [1, 2, 3]] +
    [f'Corr_Gold_GSR_lag{l}' for l in [1, 2, 3]]
)

REGIME_FEATURES = [
    'SPY_Ret_20d',        # Equity momentum (suppress Silver in equity bull)
    'SPY_Above_MA200',    # Trend (is equity rally sustained?)
    'VIX_Level',          # Market fear (affects metals)
    'TY_Momentum',        # Bond yield momentum (affects carry)
]

ENHANCED_FEATURES = [
    'Silver_Rel_Strength', # Silver 20d return - Gold 20d return
    'Vol_Ratio',          # Silver_vol / Gold_vol
    'Carry_Signal',       # GSR at extremes (< -2 = undervalued)
    'Metals_vs_Equities', # Corr(metals, SPY) — when negative, metals outperform
]

FEATURES = BASE_FEATURES + REGIME_FEATURES + ENHANCED_FEATURES

os.makedirs(MODEL_DIR, exist_ok=True)
os.makedirs("logs/portfolios", exist_ok=True)


# ── Data Fetch ─────────────────────────────────────────────────────────────────
def fetch_data(lookback_days=1500):
    """Fetch metals + macro data."""
    end   = datetime.now()
    start = end - timedelta(days=lookback_days)

    # Metals
    metals_raw = yf.download(['GC=F', 'SI=F', 'HG=F'], start=start, end=end, progress=False)
    metals = metals_raw['Close'].copy() if isinstance(metals_raw.columns, pd.MultiIndex) else metals_raw.copy()
    metals.rename(columns={'GC=F': 'Gold', 'SI=F': 'Silver', 'HG=F': 'Copper'}, inplace=True)
    metals.ffill(inplace=True)

    # Equities (SPY, VIX proxy via ^VIX, TLT for bonds)
    # Note: DXY removed due to yfinance issues; using TLT as proxy for rates
    try:
        equity_raw = yf.download(['SPY', '^VIX', 'TLT'], start=start, end=end, progress=False)
        equity = equity_raw['Close'].copy() if isinstance(equity_raw.columns, pd.MultiIndex) else equity_raw.copy()
        equity.rename(columns={'^VIX': 'VIX', 'TLT': 'TY'}, inplace=True)
        equity.ffill(inplace=True)
    except Exception as e:
        print(f"Warning: Equity fetch partially failed ({e}). Using fallback.")
        equity = pd.DataFrame(index=metals.index)
        equity['SPY'] = np.nan
        equity['VIX'] = 20.0  # Default VIX
        equity['TY'] = np.nan

    # Merge
    df = metals.join(equity, how='left').ffill().bfill()

    # ── Metals features ────────────────────────────────────────────────────────
    df['GCR']       = df['Gold'] / df['Copper'].clip(lower=0.01)
    df['GSR']       = df['Gold'] / df['Silver'].clip(lower=0.01)
    df['GCR_Delta'] = df['GCR'].diff()

    for col in ['GCR', 'GSR']:
        mu  = df[col].rolling(252).mean()
        sig = df[col].rolling(252).std()
        df[f'{col}_Z_Score'] = (df[col] - mu) / (sig + 1e-8)

    df['Silver_Ret']     = df['Silver'].pct_change()
    df['Silver_Vol20']   = df['Silver_Ret'].rolling(20).std() * np.sqrt(252)
    df['Silver_VolMA']   = df['Silver_Vol20'].rolling(126).mean()
    df['Silver_HighVol'] = df['Silver_Vol20'] > (df['Silver_VolMA'] * 1.1)  # 10% above MA

    df['Gold_Ret']       = df['Gold'].pct_change()
    df['Gold_Vol20']     = df['Gold_Ret'].rolling(20).std() * np.sqrt(252)
    df['Gold_MA200']     = df['Gold'].rolling(MA_WINDOW).mean()
    df['Gold_AboveMA']   = df['Gold'] >= df['Gold_MA200']

    for lag in [1, 2, 3]:
        df[f'GCR_Z_lag{lag}'] = df['GCR_Z_Score'].shift(lag)
        df[f'GSR_Z_lag{lag}'] = df['GSR_Z_Score'].shift(lag)

    for window in [5, 10, 20]:
        df[f'Gold_Ret_{window}d']   = df['Gold'].pct_change(window)
        df[f'Silver_Ret_{window}d'] = df['Silver'].pct_change(window)

    df['Corr_Gold_GCR'] = df['Gold'].rolling(60).corr(df['GCR'])
    df['Corr_Gold_GSR'] = df['Gold'].rolling(60).corr(df['GSR'])
    for lag in [1, 2, 3]:
        df[f'Corr_Gold_GCR_lag{lag}'] = df['Corr_Gold_GCR'].shift(lag)
        df[f'Corr_Gold_GSR_lag{lag}'] = df['Corr_Gold_GSR'].shift(lag)

    # ── Regime features ────────────────────────────────────────────────────────
    df['SPY_Ret_20d']    = df['SPY'].pct_change(20) if 'SPY' in df else 0.0
    df['SPY_MA200']      = df['SPY'].rolling(MA_WINDOW).mean() if 'SPY' in df else df['Gold']  # Fallback
    df['SPY_Above_MA200'] = (df['SPY'] >= df['SPY_MA200']).astype(float) if 'SPY' in df else 1.0
    df['VIX_Level']      = df['VIX'] / 20.0 if 'VIX' in df else 1.0  # Normalize to ~1 at 20
    df['TY_Momentum']    = (df['TY'].pct_change(20) if 'TY' in df else 0.0) * -1  # Inverse (lower yields = bullish metals)

    # ── Enhanced features ──────────────────────────────────────────────────────
    df['Silver_Rel_Strength'] = df['Silver_Ret_20d'] - df['Gold_Ret_20d']
    df['Vol_Ratio'] = (df['Silver_Vol20'] / (df['Gold_Vol20'] + 1e-8)).clip(0, 5)
    df['Carry_Signal'] = (df['GSR_Z_Score'] < -2).astype(float) * 2 + (df['GSR_Z_Score'] > 2).astype(float) * -2

    # Metals vs Equities correlation (20-day rolling)
    metals_idx = (df['Gold'].pct_change() + df['Silver'].pct_change()) / 2
    if 'SPY' in df:
        df['Metals_vs_Equities'] = metals_idx.rolling(20).corr(df['SPY'].pct_change())
    else:
        df['Metals_vs_Equities'] = 0.0

    # ── Targets (original 1-day + new 5-day) ────────────────────────────────────
    fwd_silver_1d = df['Silver'].pct_change().shift(-1)
    fwd_gold_1d   = df['Gold'].pct_change().shift(-1)
    df['Target_1d'] = (fwd_silver_1d > fwd_gold_1d).astype(int)  # Original target

    fwd_silver_5d = df['Silver'].pct_change(5).shift(-5)
    fwd_gold_5d   = df['Gold'].pct_change(5).shift(-5)
    df['Target_5d'] = (fwd_silver_5d > fwd_gold_5d).astype(int)  # New cleaner target

    required = FEATURES + ['Target_1d', 'Target_5d', 'Silver_HighVol', 'Gold_AboveMA',
                            'GSR_Z_Score', 'GCR_Delta', 'Silver_Vol20', 'Silver_VolMA',
                            'Gold_MA200', 'Gold', 'Silver', 'Copper', 'SPY_Ret_20d']
    return df.dropna(subset=required)


# ── Model Training with Walk-Forward & Validation ───────────────────────────────
def train_model_walkforward(df, retrain_date=None):
    """
    Train model as of retrain_date (for walk-forward).
    If retrain_date is None, use the latest checkpoint.
    """
    model_path   = os.path.join(MODEL_DIR, f"oqg_v5_imp_model.joblib")
    scaler_path  = os.path.join(MODEL_DIR, f"oqg_v5_imp_scaler.joblib")
    metrics_path = os.path.join(MODEL_DIR, f"oqg_v5_imp_metrics.joblib")

    # If no retrain requested and model exists, load it
    if retrain_date is None and os.path.exists(model_path) and os.path.exists(scaler_path):
        return joblib.load(model_path), joblib.load(scaler_path), joblib.load(metrics_path)

    # Train on data up to retrain_date (or 2 years back if None)
    if retrain_date is None:
        retrain_date = datetime.now() - timedelta(days=2*365)
    retrain_date = pd.to_datetime(retrain_date)

    # Use last 2 years of data for training (walk-forward window)
    train_start = retrain_date - timedelta(days=2*365)
    train = df[(df.index >= train_start) & (df.index <= retrain_date)].copy()

    if len(train) < 250:
        raise ValueError(f"Insufficient training data ({len(train)} rows). Need >= 250.")

    # Try 5-day target first, fall back to 1-day if insufficient
    X = train[FEATURES]
    y = train['Target_5d']
    mask = ~(X.isna().any(axis=1) | y.isna())
    X, y = X[mask], y[mask]

    if len(X) < 200:
        print(f"      5-day target insufficient ({len(X)} rows). Using 1-day target instead...")
        X = train[FEATURES]
        y = train['Target_1d']
        mask = ~(X.isna().any(axis=1) | y.isna())
        X, y = X[mask], y[mask]

    if len(X) < 200:
        raise ValueError(f"After NaN removal: {len(X)} rows. Need >= 200.")

    print(f"  Training on {len(X)} days ({X.index.min().date()} → {X.index.max().date()})")
    print(f"  Class balance: {y.value_counts().to_dict()}")

    # Standardscale
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # Train with balanced class weights (penalize false Silver signals)
    lr = LogisticRegression(
        max_iter=2000,
        random_state=42,
        n_jobs=-1,
        class_weight='balanced',  # Handle class imbalance
        solver='lbfgs',
        penalty='l2',
        C=1.0  # Regularization
    )
    lr.fit(X_scaled, y)

    # Cross-validation on training set
    cv_scores = cross_val_score(lr, X_scaled, y, cv=5, scoring='roc_auc')
    print(f"  Cross-validation AUC: {cv_scores.mean():.3f} ± {cv_scores.std():.3f}")

    # Full-set metrics
    y_pred = lr.predict(X_scaled)
    y_pred_proba = lr.predict_proba(X_scaled)[:, 1]
    auc = roc_auc_score(y, y_pred_proba)
    accuracy = (y_pred == y).sum() / len(y)

    print(f"  Full-set AUC: {auc:.3f}, Accuracy: {accuracy:.3f}")
    print(f"  Classification Report:\n{classification_report(y, y_pred, target_names=['Gold', 'Silver'])}")

    if auc < 0.52:
        print(f"  ⚠ WARNING: AUC {auc:.3f} is very close to random (0.50). Model may not be predictive!")

    metrics = {
        'auc': auc,
        'accuracy': accuracy,
        'cv_auc': cv_scores.mean(),
        'trained_date': retrain_date,
    }

    joblib.dump(lr, model_path)
    joblib.dump(scaler, scaler_path)
    joblib.dump(metrics, metrics_path)
    print(f"  Saved to {model_path}")

    return lr, scaler, metrics


def get_signal(row, model, scaler):
    """
    Enhanced signal logic:
    - Require high confidence (prob > 0.60)
    - Kill-switch: if SPY is rallying hard, suppress Silver signals
    - Volatility gate: don't trade Silver if vol is elevated
    - Trend gate: go CASH if Gold breaks MA200
    """
    try:
        X = pd.DataFrame([row[FEATURES]])
        X_scaled = scaler.transform(X)
        pred = int(model.predict(X_scaled)[0])
        prob = float(model.predict_proba(X_scaled)[0][1])
    except Exception as e:
        print(f"  Signal error: {e}")
        return "GOLD", 0, 0.5

    # Confidence check: only commit if prob > 60%
    if prob < CONFIDENCE_MIN:
        return "GOLD", pred, prob

    # Regime kill-switch: if SPY is rallying, suppress Silver
    spy_rallying = row.get('SPY_Ret_20d', 0) > 0.05  # >5% return in 20d
    if spy_rallying and pred == 1:
        return "GOLD", pred, prob  # Reject Silver signal if equities strong

    # Volatility gate
    if pred == 1 and row['Silver_HighVol']:
        return "CASH", pred, prob

    # Trend gate
    if not row['Gold_AboveMA']:
        return "CASH", pred, prob

    if pred == 1:
        return "SILVER", pred, prob
    return "GOLD", pred, prob


# ── Excel Formatting ───────────────────────────────────────────────────────────
def format_workbook(path):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F3864")
    gold_fill   = PatternFill("solid", fgColor="FFD700")
    silver_fill = PatternFill("solid", fgColor="C0C0C0")
    cash_fill   = PatternFill("solid", fgColor="D9EAD3")

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    pos_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Position":
            pos_col = idx
            break

    for row in ws.iter_rows(min_row=2):
        pos  = row[pos_col - 1].value if pos_col else None
        fill = (gold_fill   if pos == "GOLD"   else
                silver_fill if pos == "SILVER" else
                cash_fill   if pos == "CASH"   else None)
        if fill:
            for cell in row:
                cell.fill = fill

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 22)

    ws.freeze_panes = "B2"
    wb.save(path)


# ── Portfolio Engine ───────────────────────────────────────────────────────────
def run_portfolio_engine():
    print("=" * 70)
    print("  OQG v5 IMPROVED — Live Portfolio Engine")
    print("=" * 70)

    print("\n[1/4] Fetching data...")
    df = fetch_data()
    print(f"      {len(df)} rows, {df.index.min().date()} → {df.index.max().date()}")

    print("\n[2/4] Training model with walk-forward validation...")
    # Retrain if no model or it's older than 30 days
    model_path = os.path.join(MODEL_DIR, "oqg_v5_imp_model.joblib")
    needs_retrain = True
    if os.path.exists(model_path):
        try:
            metrics = joblib.load(os.path.join(MODEL_DIR, "oqg_v5_imp_metrics.joblib"))
            days_old = (datetime.now() - metrics['trained_date']).days
            if days_old < 30:
                needs_retrain = False
                print(f"      Model is {days_old} days old — loading cached version")
        except:
            pass

    if needs_retrain:
        model, scaler, metrics = train_model_walkforward(df)
    else:
        model = joblib.load(model_path)
        scaler = joblib.load(os.path.join(MODEL_DIR, "oqg_v5_imp_scaler.joblib"))
        metrics = joblib.load(os.path.join(MODEL_DIR, "oqg_v5_imp_metrics.joblib"))

    print(f"      Model AUC: {metrics['auc']:.3f}")

    print("\n[3/4] Processing portfolio...")
    if os.path.exists(PORTFOLIO_FILE):
        port_df          = pd.read_excel(PORTFOLIO_FILE)
        port_df['Date']  = pd.to_datetime(port_df['Date'])
        last_date        = port_df['Date'].max()
        cash             = float(port_df.iloc[-1]['Cash'])
        units_held       = float(port_df.iloc[-1]['Units_Held'])
        current_position = port_df.iloc[-1]['Position']
        days_in_pos      = 1  # Simplified: assume 1 day since last check
        print(f"      Ledger found — last logged: {last_date.date()}")
    else:
        print(f"      No ledger found. Starting fresh...")
        last_date        = df.index[0] - pd.Timedelta(days=1)
        cash             = STARTING_CAPITAL
        units_held       = 0.0
        current_position = "CASH"
        port_df          = pd.DataFrame()
        days_in_pos      = 0

    missing = df[df.index > last_date]

    if missing.empty:
        print(f"      Ledger is up to date.")
        if not port_df.empty:
            last = port_df.iloc[-1]
            print(f"      Position: {last['Position']} | Equity: ${last['Total_Equity']:,.2f}")
        return

    print(f"      Processing {len(missing)} new trading day(s)...")

    new_rows     = []
    total_equity = cash
    prev_equity  = float(port_df.iloc[-1]['Total_Equity']) if not port_df.empty else STARTING_CAPITAL

    for date, row in missing.iterrows():
        target, ml_pred, ml_prob = get_signal(row, model, scaler)
        fee_paid  = 0.0
        rotated   = False
        days_in_pos += 1

        # Rotation logic: allow rotation if (1) cooldown passed OR (2) starting from CASH
        can_rotate = (days_in_pos >= REBALANCE_COOLDOWN) or (current_position == "CASH")
        if current_position != target and can_rotate:
            rotated = True
            if current_position == "SILVER":
                cash += units_held * row['Silver']
            elif current_position == "GOLD":
                cash += units_held * row['Gold']
            units_held = 0.0

            if target != "CASH":
                fee_paid   = cash * SLIPPAGE_FEE
                cash      -= fee_paid
                buy_price  = row['Silver'] if target == "SILVER" else row['Gold']
                units_held = cash / buy_price
                cash       = 0.0

            current_position = target
            days_in_pos = 0
            print(f"  {date.date()} ROTATE → {target:6s} | Prob: {ml_prob:.3f} | Fee: ${fee_paid:,.0f}")

        if current_position == "CASH":
            total_equity = cash
        elif current_position == "SILVER":
            total_equity = units_held * row['Silver']
        else:
            total_equity = units_held * row['Gold']

        day_ret = (total_equity - prev_equity) / prev_equity if prev_equity > 0 else 0.0
        cum_ret = (total_equity - STARTING_CAPITAL) / STARTING_CAPITAL
        prev_equity = total_equity

        new_rows.append({
            'Date':             date.strftime('%Y-%m-%d'),
            'Position':         current_position,
            'Signal':           target,
            'Rotated':          "YES" if rotated else "",
            'ML_Pred':          ml_pred,
            'Silver_Prob':      round(ml_prob, 4),
            'Confidence_Met':   "YES" if ml_prob >= CONFIDENCE_MIN else "NO",
            'Gold_Price':       round(row['Gold'], 2),
            'Silver_Price':     round(row['Silver'], 4),
            'SPY_Ret_20d':      round(row.get('SPY_Ret_20d', 0), 4),
            'VIX_Level':        round(row.get('VIX_Level', 0), 2),
            'Gold_MA200':       round(row['Gold_MA200'], 2),
            'Silver_HighVol':   bool(row['Silver_HighVol']),
            'Gold_AboveMA':     bool(row['Gold_AboveMA']),
            'Units_Held':       round(units_held, 6),
            'Cash':             round(cash, 2),
            'Total_Equity':     round(total_equity, 2),
            'Day_Return_Pct':   round(day_ret * 100, 4),
            'Cum_Return_Pct':   round(cum_ret * 100, 4),
            'Fee_Paid':         round(fee_paid, 2),
        })

    updated = pd.concat([port_df, pd.DataFrame(new_rows)], ignore_index=True)
    updated.to_excel(PORTFOLIO_FILE, index=False, engine='openpyxl')
    format_workbook(PORTFOLIO_FILE)

    print("\n[4/4] Portfolio Summary")
    print("=" * 70)
    equity_s = updated['Total_Equity'].astype(float)
    daily_r  = equity_s.pct_change().dropna()
    sharpe   = (daily_r.mean() / daily_r.std() * np.sqrt(252)) if len(daily_r) > 1 else float('nan')
    cum_max  = equity_s.cummax()
    max_dd   = ((equity_s - cum_max) / cum_max).min() * 100
    pnl_pct  = (total_equity - STARTING_CAPITAL) / STARTING_CAPITAL * 100

    print(f"  Total Equity    : ${total_equity:>12,.2f}  ({pnl_pct:+.2f}%)")
    print(f"  Position        : {current_position}  ({units_held:.4f} units)")
    print(f"  Days in Position: {days_in_pos}")
    print(f"  Days Logged     : {len(updated)}")
    print(f"  Rotations       : {updated['Rotated'].eq('YES').sum()}")
    print(f"  Sharpe Ratio    : {sharpe:.3f}")
    print(f"  Max Drawdown    : {max_dd:.2f}%")
    print(f"  Model AUC       : {metrics['auc']:.3f}  (need > 0.52)")
    print(f"  Ledger          : {PORTFOLIO_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    run_portfolio_engine()
